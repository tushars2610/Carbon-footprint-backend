# scripts/ingest.py
from google.cloud import storage
import logging
from typing import Dict, Any
from .utils import setup_logging, gcs_write_json, gcs_read_file, gcs_list_files
import os
import json
from docx import Document
import PyPDF2
import pytesseract
from PIL import Image
from openpyxl import load_workbook
import nltk
from bs4 import BeautifulSoup
from tqdm import tqdm
from dotenv import load_dotenv
import tempfile
from google.oauth2 import service_account

# Load environment variables
load_dotenv()
GCS_BUCKET_NAME = os.getenv("GCS_BUCKET_NAME")

def get_gcs_client():
    """Get Google Cloud Storage client with proper credential handling"""
    
    # Check if credentials are provided as JSON string (for Render/production)
    credentials_json = os.getenv('GOOGLE_APPLICATION_CREDENTIALS_JSON')
    if credentials_json:
        try:
            credentials_info = json.loads(credentials_json)
            credentials = service_account.Credentials.from_service_account_info(credentials_info)
            return storage.Client(credentials=credentials)
        except json.JSONDecodeError as e:
            logging.error(f"Failed to parse GOOGLE_APPLICATION_CREDENTIALS_JSON: {e}")
            raise
    
    # Check if credentials file path is provided (for local development)
    credentials_file = os.getenv('GOOGLE_APPLICATION_CREDENTIALS')
    if credentials_file and os.path.exists(credentials_file):
        return storage.Client.from_service_account_json(credentials_file)
    
    # Try default credentials (when running on GCP)
    try:
        return storage.Client()
    except Exception as e:
        logging.error(f"Failed to initialize GCS client with default credentials: {e}")
        raise Exception(
            "Unable to authenticate with Google Cloud Storage. "
            "Please set either GOOGLE_APPLICATION_CREDENTIALS_JSON (JSON string) "
            "or GOOGLE_APPLICATION_CREDENTIALS (file path) environment variable."
        )

class DocumentIngester:
    def __init__(self, domain: str, config: Dict[str, Any], base_dir: str):
        self.domain = domain
        self.input_dir = f"{base_dir}/docs/{domain}"
        self.output_dir = f"{base_dir}/data/{domain}/json"
        self.config = config
        self.supported_extensions = {'.txt', '.docx', '.pdf', '.jpg', '.png', '.xlsx'}
        self.client = self.client = self.client = get_gcs_client()  # Use the helper function
        self.bucket = self.client.bucket(GCS_BUCKET_NAME)
        setup_logging(config)

    def normalize_text(self, text: str) -> str:
        """Normalize text."""
        if not text:
            return ""
        text = ' '.join(text.split())
        if '<' in text and '>' in text:
            soup = BeautifulSoup(text, 'html.parser')
            text = soup.get_text(separator=' ', strip=True)
        return text.strip()

    def extract_text_from_txt(self, blob_path: str) -> str:
        try:
            content = gcs_read_file(blob_path)
            if not content:
                logging.warning(f"Empty content in {blob_path}")
                return ""
            return self.normalize_text(content)
        except Exception as e:
            logging.error(f"Error reading {blob_path}: {e}")
            return ""

    def extract_text_from_docx(self, blob_path: str) -> str:
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as temp_file:
                bucket_name, blob_name = self.parse_gcs_path(blob_path)
                blob = self.bucket.blob(blob_name)
                blob.download_to_filename(temp_file.name)
                doc = Document(temp_file.name)
                text = '\n'.join([para.text for para in doc.paragraphs if para.text.strip()])
            os.unlink(temp_file.name)
            if not text:
                logging.warning(f"Empty content extracted from {blob_path}")
                return ""
            return self.normalize_text(text)
        except Exception as e:
            logging.error(f"Error reading {blob_path}: {e}")
            return ""

    def extract_text_from_pdf(self, blob_path: str) -> str:
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
                bucket_name, blob_name = self.parse_gcs_path(blob_path)
                blob = self.bucket.blob(blob_name)
                blob.download_to_filename(temp_file.name)
                with open(temp_file.name, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    text = ""
                    for page in reader.pages:
                        page_text = page.extract_text()
                        if page_text and len(page_text.strip()) > 10:
                            text += page_text + '\n'
                        else:
                            logging.info(f"Performing OCR on {blob_path}")
                            images = page.images
                            for image in images:
                                img = Image.open(image.data)
                                text += pytesseract.image_to_string(img) + '\n'
            os.unlink(temp_file.name)
            if not text:
                logging.warning(f"Empty content extracted from {blob_path}")
                return ""
            return self.normalize_text(text)
        except Exception as e:
            logging.error(f"Error reading {blob_path}: {e}")
            return ""

    def extract_text_from_image(self, blob_path: str) -> str:
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as temp_file:
                bucket_name, blob_name = self.parse_gcs_path(blob_path)
                blob = self.bucket.blob(blob_name)
                blob.download_to_filename(temp_file.name)
                img = Image.open(temp_file.name)
                text = pytesseract.image_to_string(img)
            os.unlink(temp_file.name)
            if not text:
                logging.warning(f"Empty content extracted from {blob_path}")
                return ""
            return self.normalize_text(text)
        except Exception as e:
            logging.error(f"Error processing image {blob_path}: {e}")
            return ""

    def extract_data_from_xlsx(self, blob_path: str) -> list:
        temp_file_path = None
        try:
            # Parse GCS path
            bucket_name, blob_name = self.parse_gcs_path(blob_path)
            
            # Get the correct bucket and blob
            client = storage.Client()  # or use self.client if you have it
            bucket = client.bucket(bucket_name)
            blob = bucket.blob(blob_name)
            
            # Check if blob exists
            if not blob.exists():
                logging.error(f"Blob does not exist: {blob_path}")
                return []
            
            # Create temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                temp_file_path = temp_file.name
                
            # Download blob to temporary file
            logging.info(f"Downloading {blob_path} to {temp_file_path}")
            blob.download_to_filename(temp_file_path)
            
            # Verify file was downloaded and has content
            if not os.path.exists(temp_file_path) or os.path.getsize(temp_file_path) == 0:
                logging.error(f"Downloaded file is empty or doesn't exist: {temp_file_path}")
                return []
            
            # Load workbook
            logging.info(f"Loading workbook from {temp_file_path}")
            wb = load_workbook(temp_file_path, read_only=True)
            
            if not wb.worksheets:
                logging.warning(f"No worksheets found in {blob_path}")
                return []
            
            data = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                logging.info(f"Processing sheet: {sheet_name}")
                
                # Get all rows
                rows = list(sheet.iter_rows(values_only=True))
                
                if not rows:
                    logging.info(f"Sheet {sheet_name} is empty")
                    continue
                    
                # Filter out completely empty rows
                non_empty_rows = [row for row in rows if any(cell is not None and str(cell).strip() for cell in row)]
                
                if not non_empty_rows:
                    logging.info(f"Sheet {sheet_name} has no non-empty rows")
                    continue
                
                headers = non_empty_rows[0]
                
                # Validate headers
                if not any(header is not None and str(header).strip() for header in headers):
                    logging.warning(f"Sheet {sheet_name} has no valid headers")
                    continue
                
                logging.info(f"Found {len(headers)} headers: {headers}")
                
                # Process data rows
                for i, row in enumerate(non_empty_rows[1:], start=2):
                    try:
                        row_data = {}
                        for j, (header, value) in enumerate(zip(headers, row)):
                            header_str = str(header).strip() if header is not None else f"Column_{j+1}"
                            value_str = str(value).strip() if value is not None else ""
                            row_data[header_str] = value_str
                        
                        # Only add rows that have at least one non-empty value
                        if any(v for v in row_data.values()):
                            data.append(row_data)
                            
                    except Exception as row_error:
                        logging.warning(f"Error processing row {i} in sheet {sheet_name}: {row_error}")
                        continue
            
            wb.close()
            logging.info(f"Successfully extracted {len(data)} records from {blob_path}")
            
            if not data:
                logging.warning(f"No data extracted from {blob_path}")
            
            return data
            
        except Exception as e:
            logging.error(f"Error reading {blob_path}: {str(e)}", exc_info=True)
            return []
            
        finally:
            # Clean up temporary file
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                    logging.info(f"Cleaned up temporary file: {temp_file_path}")
                except Exception as cleanup_error:
                    logging.warning(f"Failed to clean up temporary file {temp_file_path}: {cleanup_error}")

    def parse_gcs_path(self, gcs_path: str) -> tuple:
        """Parse GCS path into bucket name and blob name."""
        if not gcs_path.startswith('gs://'):
            raise ValueError(f"Invalid GCS path: {gcs_path}")
        parts = gcs_path.replace('gs://', '').split('/', 1)
        if len(parts) < 2:
            raise ValueError(f"Invalid GCS path: {gcs_path}")
        return parts[0], parts[1]

    def process_file(self, file_path: str) -> Dict[str, Any]:
        """Process a single file from GCS and return structured data."""
        file_name = file_path.rsplit('/', 1)[-1]
        ext = ('.' + file_name.rsplit('.', 1)[-1].lower()) if '.' in file_name else ''
        blob_name = file_path.replace(f"gs://{GCS_BUCKET_NAME}/", '')
        try:
            blob = self.bucket.get_blob(blob_name)
            if not blob:
                logging.error(f"Blob {blob_name} not found in GCS")
                return None
            metadata = {
                "file_name": file_name,
                "file_path": file_path.replace(f"gs://{GCS_BUCKET_NAME}/docs/{self.domain}/", ''),
                "file_type": ext,
                "size_bytes": blob.size or 0,
                "last_modified": blob.updated.timestamp() if blob.updated else 0,
                **self.config.get('metadata', {})
            }
        except Exception as e:
            logging.error(f"Error fetching metadata for {file_path}: {e}")
            return None

        logging.debug(f"Processing file: {file_path} with extension: {ext}")

        if ext not in self.supported_extensions:
            logging.warning(f"Unsupported file type: {ext} for {file_name}")
            return None

        if ext == '.txt':
            content = self.extract_text_from_txt(file_path)
        elif ext == '.docx':
            content = self.extract_text_from_docx(file_path)
        elif ext == '.pdf':
            content = self.extract_text_from_pdf(file_path)
        elif ext in {'.jpg', '.png'}:
            content = self.extract_text_from_image(file_path)
        elif ext == '.xlsx':
            content = self.extract_data_from_xlsx(file_path)
        else:
            logging.warning(f"Unexpected file type: {ext} for {file_name}")
            return None

        if not content:
            logging.warning(f"No valid content extracted from {file_path}")
            return None

        return {"metadata": metadata, "content": content}

    def ingest(self, skip_existing: bool = False) -> list:
        """Process all files in the domain's input directory in GCS."""
        processed_files = []
        files = gcs_list_files(GCS_BUCKET_NAME, f"docs/{self.domain}/")
        logging.info(f"Found {len(files)} files for domain {self.domain}")
        
        for file_path in tqdm(files, desc=f"Ingesting files for {self.domain}"):
            file_name = file_path.rsplit('/', 1)[-1]
            ext = ('.' + file_name.rsplit('.', 1)[-1].lower()) if '.' in file_name else ''
            if ext not in self.supported_extensions:
                logging.warning(f"Skipping file with unsupported extension: {file_path}")
                continue
            
            output_path = f"{self.output_dir}/{file_name.rsplit('.', 1)[0]}.json"
            if skip_existing and self.bucket.blob(output_path.replace(f"gs://{GCS_BUCKET_NAME}/", '')).exists():
                logging.info(f"Skipping existing file: {output_path}")
                continue
            
            result = self.process_file(file_path)
            if result:
                processed_files.append(result)
                gcs_write_json(result, output_path)
            else:
                logging.warning(f"Failed to process file: {file_path}")
        
        logging.info(f"Processed {len(processed_files)} valid files for domain {self.domain}")
        return processed_files